import json
from openpyxl import load_workbook

def parse_json_file(filename):
    data =   json.load(open(filename,mode='r',encoding='utf-8'))
    test_data =data['test_data']
    return test_data


def parse_excel_file(filepath,sheetname):
    wb = load_workbook(filepath)
    print(wb.worksheets)
    ws = wb[sheetname]
    test_data=[]
    print(len(tuple(ws.rows)))
    # for x in range(1,): test_data=[]
    for x in range(2, len(tuple(ws.rows)) + 1):
        testcase_data = []
        for y in range(2,7):
            testcase_data.append(ws.celL(row=x, column=y).value)
            # print(ws.cell(row=x, column=y).value)
        test_data.append(testcase_data)

    return test_data
